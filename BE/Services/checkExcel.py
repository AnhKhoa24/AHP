import json
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

def check_ahp_table_from_wb(wb, sheet_name=None, tol=1e-6):
 
    if isinstance(wb, Worksheet):
        ws = wb
    else:
        ws = wb[sheet_name] if sheet_name else wb.active

    sheet_key = ws.title

    markers = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == "*":
                markers.append(cell)

    if len(markers) < 2:
        raise ValueError("Không tìm thấy đủ 2 ô chứa '*' để xác định vùng bảng.")

    markers.sort(key=lambda c: (c.row, c.column))
    top_left = markers[0]
    bottom_right = markers[1]

    r1, c1 = top_left.row, top_left.column
    r2, c2 = bottom_right.row, bottom_right.column

    labels = []
    for col_idx in range(c1 + 1, c2):
        cell = ws.cell(row=r1, column=col_idx)
        if cell.value is None or str(cell.value).strip() == "":
            labels.append("null")
        else:
            labels.append(str(cell.value))

    num_rows = (r2 - 1) - (r1 + 1) + 1 
    num_cols = (c2 - 1) - (c1 + 1) + 1 

    if num_rows != num_cols:
        raise ValueError(f"Vùng ma trận không có kích thước vuông (rows={num_rows}, cols={num_cols})")

    n = num_rows

    errors = []
    for i in range(n):
        for j in range(n):
            cell_ij = ws.cell(row=r1 + 1 + i, column=c1 + 1 + j)
            try:
                v_ij = float(cell_ij.value)
            except (TypeError, ValueError):
                v_ij = None

            if i == j:
                if v_ij is None or abs(v_ij - 1.0) > tol:
                    errors.append({"x": i, "y": j})
            else:
                cell_ji = ws.cell(row=r1 + 1 + j, column=c1 + 1 + i)
                try:
                    v_ji = float(cell_ji.value)
                except (TypeError, ValueError):
                    v_ji = None

                if v_ij is None or v_ji is None or abs(v_ij * v_ji - 1.0) > tol:
                    errors.append({"x": i, "y": j})
    return {
            "label": labels,
            "error": errors
    }
