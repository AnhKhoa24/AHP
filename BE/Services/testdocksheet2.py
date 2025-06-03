import openpyxl
import json
from fractions import Fraction

def to_fraction_str(value):
    try:
        frac = Fraction(value).limit_denominator()
    except (ValueError, TypeError):
        return "0"

    if frac.denominator == 1:
        return str(frac.numerator)
    else:
        return f"{frac.numerator}/{frac.denominator}"


def extract_sheet1(path_to_excel):

    wb = openpyxl.load_workbook(path_to_excel, data_only=True)
    if len(wb.worksheets) < 1:
        raise RuntimeError("File Excel không có sheet1.")
    sheet = wb.worksheets[0]

    marker_rows = []
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if any(cell == "*" for cell in row if cell is not None):
            marker_rows.append(idx)

    if len(marker_rows) < 2:
        raise RuntimeError("Không tìm thấy đủ 2 dấu '*' trong Sheet1.")

    header_row_idx = marker_rows[0]
    footer_row_idx = marker_rows[1]

    header_cells = list(sheet[header_row_idx])
    labels = []
    label_column_indices = []
    for col_idx, cell in enumerate(header_cells, start=1):
        val = cell.value
        if val not in (None, "", "*"):
            labels.append(str(val))
            label_column_indices.append(col_idx)

    # 3) Đọc ma trận: từ header_row_idx+1 đến footer_row_idx-1
    matrix = []
    for r in range(header_row_idx + 1, footer_row_idx):
        # Nếu hàng trống, bỏ qua
        row_vals = [cell.value for cell in sheet[r]]
        if all(v in (None, "") for v in row_vals):
            continue

        row_fraction = []
        for c in label_column_indices:
            raw = sheet.cell(row=r, column=c).value
            frac_str = to_fraction_str(raw)
            row_fraction.append(frac_str)
        matrix.append(row_fraction)

    return {
        "label": labels,
        "matran": matrix
    }


def extract_sheet2(path_to_excel):
    """
    Đọc Sheet2 (index = 1), với mỗi cặp dấu '*':
      - Tìm "text" bằng cách quét ngược lên phía trên, rồi chỉ giữ phần sau dấu ':' (nếu có).
      - Tìm các cột header (bỏ qua ô '*' và rỗng), sau đó đọc từng hàng bên dưới
        thành chuỗi phân số.
    Trả về list các dict:
      [
        {
          "text": "...",
          "matran": [
             ["1/3", "2", ...],
             ...
          ]
        },
        ...
      ]
    """
    wb = openpyxl.load_workbook(path_to_excel, data_only=True)
    if len(wb.worksheets) < 2:
        raise RuntimeError("File Excel không có sheet2.")
    sheet = wb.worksheets[1]

    # 1) Tìm tất cả hàng chứa dấu '*'
    marker_rows = []
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if any(cell == "*" for cell in row if cell is not None):
            marker_rows.append(idx)

    if len(marker_rows) % 2 != 0:
        raise RuntimeError("Số lượng dấu '*' trên Sheet2 không phải là số chẵn.")

    results = []

    # 2) Với mỗi cặp (start, end)
    for i in range(0, len(marker_rows), 2):
        header_row_idx = marker_rows[i]
        footer_row_idx = marker_rows[i+1]

        # 2.1) Quét ngược lên để tìm "text"
        text_row_idx = header_row_idx - 1
        text_value = ""
        while text_row_idx >= 1:
            # Lấy toàn bộ ô ở hàng đó, chỉ quan tâm giá trị khác None/""
            row_vals = [cell.value for cell in sheet[text_row_idx] if cell is not None]
            non_empty = [v for v in row_vals if v not in (None, "")]
            if non_empty:
                text_value = str(non_empty[0])
                break
            text_row_idx -= 1

        # Nếu có dấu ':' thì chỉ lấy phần sau nó
        if ":" in text_value:
            text_value = text_value.split(":", 1)[1].strip()

        # 2.2) Xác định cột dữ liệu dựa vào header_row_idx (bỏ qua '*' và rỗng)
        header_cells = list(sheet[header_row_idx])
        data_column_indices = []
        for col_idx, cell in enumerate(header_cells, start=1):
            if cell.value not in (None, "", "*"):
                data_column_indices.append(col_idx)

        # 2.3) Đọc ma trận từ header_row_idx+1 đến footer_row_idx-1
        matrix = []
        for r in range(header_row_idx + 1, footer_row_idx):
            row_vals = [cell.value for cell in sheet[r]]
            if all(v in (None, "") for v in row_vals):
                continue

            row_fraction = []
            for c in data_column_indices:
                raw = sheet.cell(row=r, column=c).value
                frac_str = to_fraction_str(raw)
                row_fraction.append(frac_str)
            matrix.append(row_fraction)

        results.append({
            "text": text_value,
            "matran": matrix
        })

    return results


path = r"C:\Users\anhkh\Downloads\AHP (1).xlsx"

sheet1_data = extract_sheet1(path)

sheet2_data = extract_sheet2(path)

final_output = {
        "sheet1": sheet1_data,
        "sheet2": sheet2_data
    }

print(json.dumps(final_output, ensure_ascii=False, indent=4))
