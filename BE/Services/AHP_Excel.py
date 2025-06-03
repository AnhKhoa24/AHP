from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
import numpy as np

def copytb(ws, start, lengh, start_newtb, border_all):
    for i in range(0, lengh, 1): 
        for j in range(0, lengh, 1):
            col = get_column_letter(j + 1)
            cell = ws[f"{col}{start_newtb+i}"]
            if i == 0 and j == 0:
                cell.value = ""
            else:
                cell.value = f"={col}{i+start}"
            cell.font = Font(bold=False, color="000000", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if i == 0 or j == 0:
                cell.fill = PatternFill("solid", fgColor="5dade2")  
            cell.border = border_all

def colSum(ws, start, lengh, start_newcol, border_all):
    cell = ws[f"A{start_newcol}"]
    cell.value = "SUM"
    cell.font = Font(bold=False, color="000000", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor="f0b27a")  
    cell.border = border_all
    for i in range (0, lengh, 1):
        col = get_column_letter(i + 2)
        cell = ws[f"{col}{start_newcol}"]
        cell.value = f"=SUM({col}{start}:{col}{start+lengh})"
        cell.font = Font(bold=False, color="000000", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="f0b27a")  
        cell.border = border_all

def generate_label(ws, start, lengh, content):
    end_col_letter = get_column_letter(lengh + 1)
    merge_range = f"A{start}:{end_col_letter}{start}"
    ws.merge_cells(merge_range)

    cell = ws[f'A{start}']
    cell.value = content
    cell.font = Font(bold=False, color="FFFFFF", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor="4F81BD")

def generate_tieuchiFull(ws, start, tieuchis, border_all):
    ws[f"A{start}"].fill = PatternFill("solid", fgColor="5dade2")
    ws[f"A{start}"].border = border_all
    for i in range(len(tieuchis)):
        col_letter = get_column_letter(i + 2) 
        cell = ws[f"{col_letter}{start}"]
        cell.value = tieuchis[i]
        cell.font = Font(bold=False, color="000000", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="5dade2")  
        cell.border = border_all
    for i in range(len(tieuchis)):
        cell = ws[f"A{start+i+1}"]
        cell.value = tieuchis[i]
        cell.font = Font(bold=False, color="000000", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="5dade2")
        cell.border = border_all

def generate_chuanhoatb(ws, start, lengh, start_sum_col, start_newtb, border_all):
    for i in range(0, lengh, 1): 
        for j in range(0, lengh, 1):
            col = get_column_letter(i + 2)
            cell = ws[f"{col}{j+start_newtb}"]
            cell.value = f"={col}{j+start}/{col}{start_sum_col}"
            cell.font = Font(bold=False, color="000000", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_all

def generate_col_CW(ws, start, lengh, start_row, start_col, border_all):
    ws[f"{start_col}{start_row}"].fill = PatternFill("solid", fgColor="f4d03f")
    ws[f"{start_col}{start_row}"].border = border_all
    ws[f"{start_col}{start_row}"].value = "Criteria Weight"

    end_quet = get_column_letter(lengh+1)
    for i in range (0, lengh, 1):
        cell = ws[f"{start_col}{i+start_row+1}"]
        cell.value = f"=AVERAGE(B{i+start}:{end_quet}{i+start})"
        cell.font = Font(bold=False, color="000000", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="f4d03f")
        cell.border = border_all

def generate_nhanCW(ws, start, lengh, start_cw_row, start_cw_col, start_newtb, border_all):
    for i in range(0, lengh, 1):
        for j in range(0, lengh, 1):
            col = get_column_letter(i+2)
            cell = ws[f"{col}{j+start_newtb}"]
            cell.value = f"={col}{j+start} * {start_cw_col}{start_cw_row+i}"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_all
    return start_cw_row+lengh, get_column_letter(lengh+2)

def generate_ws(ws, start, lengh, start_new_col, start_new_row, border_all):
    ws[f"{start_new_col}{start_new_row}"].fill = PatternFill("solid", fgColor="f4d03f")
    ws[f"{start_new_col}{start_new_row}"].border = border_all
    ws[f"{start_new_col}{start_new_row}"].value = "Weighted Sum Value"
    for i in range(0, lengh, 1):
        cell = ws[f"{start_new_col}{start_new_row+i+1}"]
        cell.value = f"=SUM(B{start+i}:{get_column_letter(lengh+1)}{start+i})"
        cell.font = Font(bold=False, color="000000", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="f4d03f")
        cell.border = border_all
    col_idx = column_index_from_string(start_new_col) 
    return start_new_row+lengh, get_column_letter(col_idx+1)

def generate_CV(ws, start_col, start_row, lengh, new_col, new_row, border_all):
    ws[f"{new_col}{new_row}"].fill = PatternFill("solid", fgColor="f4d03f")
    ws[f"{new_col}{new_row}"].border = border_all
    ws[f"{new_col}{new_row}"].value = "Consistery Vector"

    start_col_2 = get_column_letter(column_index_from_string(start_col) + 1)
    for i in range(0, lengh, 1):
        cell = ws[f"{new_col}{new_row+i+1}"]
        cell.value = f"={start_col}{start_row+i}/{start_col_2}{start_row+i}"
        cell.font = Font(bold=False, color="000000", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="f4d03f")
        cell.border = border_all

def write_summary(ws, cv_end_row, col_letter, n, border_all):
  
    prev_col_letter = get_column_letter(column_index_from_string(col_letter) - 1)
    row = cv_end_row + 2

    # Lambda max
    cell = ws[f"{prev_col_letter}{row}"]
    cell.value = "Lamda max"
    cell.font = Font(bold=True, color="000000", size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    cell = ws[f"{col_letter}{row}"]
    cv_start_row = cv_end_row - n + 1
    cell.value = f"=AVERAGE({col_letter}{cv_start_row}:{col_letter}{cv_end_row})"
    cell.font = Font(bold=False, color="000000", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_all

    # CI
    row += 1
    cell = ws[f"{prev_col_letter}{row}"]
    cell.value = "CI"
    cell.font = Font(bold=True, color="000000", size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    cell = ws[f"{col_letter}{row}"]
    lamda_cell = f"{col_letter}{row-1}"
    cell.value = f"=({lamda_cell}-{n})/({n}-1)"
    cell.font = Font(bold=False, color="000000", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_all

    # CR
    row += 1
    cell = ws[f"{prev_col_letter}{row}"]
    cell.value = "CR"
    cell.font = Font(bold=True, color="000000", size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    cell = ws[f"{col_letter}{row}"]
    ci_cell = f"{col_letter}{row-1}"
    RI_dict = {
        1: 0.00, 2: 0.00, 3: 0.58, 4: 0.90, 5: 1.12,
        6: 1.24, 7: 1.32, 8: 1.41, 9: 1.45, 10: 1.49
    }
    RI = RI_dict.get(n, 1.49)
    cell.value = f"={ci_cell}/{RI}"
    cell.font = Font(bold=False, color="000000", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_all

def ahpExcel(tieuchis, matran):
    n = matran.shape[0]
    wb = Workbook()
    ws = wb.active
    ws.title = "Tính trọng số từng tiêu chí"

    end_col_letter = get_column_letter(len(tieuchis) + 1)
    merge_range = f"A1:{end_col_letter}1"
    ws.merge_cells(merge_range)

    cell = ws['A1']
    cell.value = "Xây dựng ma trận so sánh cặp cho từng tiêu chí"
    cell.font = Font(bold=False, color="FFFFFF", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor="4F81BD")  

    thin = Side(border_style="thin", color="000000")
    border_all = Border(top=thin, left=thin, right=thin, bottom=thin)

    bang1 = 3 
    generate_tieuchiFull(ws, bang1, tieuchis, border_all)
    cell = ws[f'A{bang1}']
    cell.value = "*"
    cell.font = Font(bold=False, color="5dade2", size=11)

    cell = ws[f'{get_column_letter(len(tieuchis) + 2)}{len(tieuchis) + bang1+1}']
    cell.value = "*"
    cell.font = Font(bold=False, color="FFFFFF", size=11)

    for i in range(n):
        for j in range(n):
            cell = ws.cell(row=bang1 + 1 + i, column=2 + j)
            if i == j:
                value = 1.0
            elif i < j:
                value = matran[i][j]
            else:
                col_letter_top = get_column_letter(2 + i)
                row_top = bang1 + 1 + j
                value = f"=1/{col_letter_top}{row_top}"

            cell.value = value
            cell.border = border_all
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # cell.number_format = "??/??" 
             
    tieude = bang1+n+2    
    generate_label(ws, tieude, n, "Tính sum cho từng cột")

    currentcol = tieude+2
    start_bangtinhsum = currentcol
    copytb(ws, bang1, n+1, start_bangtinhsum, border_all)

    currentcol += n+1
    start_col_sum = currentcol
    colSum(ws, bang1+1, n, start_col_sum, border_all)

    currentcol += 2
    generate_label(ws, currentcol, n, "Chuẩn hóa ma trận so sánh cặp")

    currentcol += 2
    generate_tieuchiFull(ws, currentcol, tieuchis, border_all)
    start_tb_chuanhoa = currentcol+1
    generate_chuanhoatb(ws, start_bangtinhsum+1, n, start_col_sum, start_tb_chuanhoa, border_all)

    start_tb_chuanhoa2 = start_tb_chuanhoa + n + 1
    generate_label(ws, start_tb_chuanhoa2, n, "Tính trọng số cho các tiêu chí tính theo từng hàng")
    start_tb_chuanhoa2 += 2
    copytb(ws, start_tb_chuanhoa-1, n+1, start_tb_chuanhoa2, border_all)

    danhdau_tb_ch = start_tb_chuanhoa2+1
    generate_col_CW(ws, danhdau_tb_ch, n, start_tb_chuanhoa2, get_column_letter(n+2), border_all)
    row_Cw = start_tb_chuanhoa2+1
    col_Cw = get_column_letter(n+2)

    start_tb_chuanhoa2 += n+2
    generate_label(ws, start_tb_chuanhoa2, n, "Sử dụng trọng số của các tiêu chí và ma trận so sánh cặp để tính tỷ số nhất quán CR")
    start_tb_chuanhoa2 += 2

    generate_tieuchiFull(ws, start_tb_chuanhoa2, tieuchis, border_all)

    start_tb_chuanhoa2 += 1
    end_row, end_col = generate_nhanCW(ws, bang1+ 1, n, row_Cw, col_Cw, start_tb_chuanhoa2 , border_all)

    end_row_gws, end_col_gws = generate_ws(ws,start_tb_chuanhoa2,n,end_col,start_tb_chuanhoa2-1, border_all )
    generate_col_CW(ws, danhdau_tb_ch, n, start_tb_chuanhoa2-1, end_col_gws, border_all)

    generate_CV(ws, end_col, start_tb_chuanhoa2, n, get_column_letter(column_index_from_string(end_col_gws) + 1), start_tb_chuanhoa2-1, border_all)


    cv_col_letter = get_column_letter(column_index_from_string(end_col_gws) + 1)
    summary_start_row = start_tb_chuanhoa2 + n -1
    write_summary(ws, summary_start_row, cv_col_letter, n, border_all)


    ####Sheet2
    ws2 = wb.create_sheet(title="Tính trọng số từng phương án")  



    wb.save("AHP2.xlsx")
    # filename = "AHP.xlsx"   
    # wb.save(filename)
    # return filename


arr = ["Địa điểm","Điểm đầu vào", "Học phí", "Chất lượng đào tạo", "Cơ sở vật chất"]
matrix = np.array([
    [1, 2, 4, 5, 1],
    [1/2, 1, 2, 3, 2],
    [1/4, 1/2, 1, 1/3, 3],
    [1/5, 1/3, 3, 1, 1/5],
    [1, 1/2, 1/3, 5, 1],
])

arr_tc = [
    {"ten": "Đại học A", "kihieu": "A"},
    {"ten": "Đại học B", "kihieu": "B"},
    {"ten": "Đại học C", "kihieu": "C"},
    {"ten": "Đại học D", "kihieu": "D"},
]

ahpExcel(arr, matrix)