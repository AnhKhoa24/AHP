from openpyxl import load_workbook
from typing import Tuple, List, Union, BinaryIO
from io import BytesIO

def extract_ahp_matrix_by_marker_safe(file: Union[bytes, BinaryIO], marker: str = "*") -> Tuple[List[str], List[List[float]]]:
    if isinstance(file, bytes):
        file = BytesIO(file)

    wb = load_workbook(file, data_only=False)
    ws = wb.active

    # Tìm vị trí 2 marker '*'
    marker_cells = []
    for row in ws.iter_rows(min_row=1, max_row=100, max_col=50):
        for cell in row:
            if str(cell.value).strip() == marker:
                marker_cells.append(cell)

    if len(marker_cells) != 2:
        raise ValueError("Không tìm đúng 2 marker để xác định bảng.")

    start_cell, end_cell = marker_cells[0], marker_cells[1]
    start_row, start_col = start_cell.row, start_cell.column
    end_row, end_col = end_cell.row, end_cell.column

    # Đọc tiêu chí ở hàng đầu
    criteria = []
    for col in range(start_col + 1, end_col):  # KHÔNG lấy cột cuối (marker)
        val = ws.cell(row=start_row, column=col).value
        if val is not None:
            criteria.append(str(val).strip())

    # Đọc ma trận so sánh
    matrix = []
    for row in range(start_row + 1, end_row):  # KHÔNG lấy dòng cuối (marker)
        row_data = []
        for col in range(start_col + 1, end_col):  # KHÔNG lấy cột cuối (marker)
            cell = ws.cell(row=row, column=col)
            val = cell.value
            try:
                if isinstance(val, (int, float)):
                    row_data.append(float(val))
                elif isinstance(val, str):
                    val = val.strip()
                    # Hỗ trợ đọc phân số như '1/3'
                    if '/' in val and not val.startswith('='):
                        num, denom = val.split('/')
                        row_data.append(round(float(num) / float(denom), 6))
                    # Hỗ trợ công thức Excel kiểu '=1/A1'
                    elif val.startswith('=1/'):
                        ref = val[3:]
                        ref_val = ws[ref].value
                        if isinstance(ref_val, (int, float)) and ref_val != 0:
                            row_data.append(round(1.0 / ref_val, 6))
                        else:
                            row_data.append(0.0)
                    else:
                        row_data.append(0.0)
                else:
                    row_data.append(0.0)
            except:
                row_data.append(0.0)
        matrix.append(row_data)

    return criteria, matrix

from fractions import Fraction

def float_to_fraction_str(value: float, max_denominator: int = 1000) -> str:
    try:
        frac = Fraction(value).limit_denominator(max_denominator)
        return f"{frac.numerator}/{frac.denominator}" if frac.denominator != 1 else f"{frac.numerator}"
    except:
        return str(value) 
