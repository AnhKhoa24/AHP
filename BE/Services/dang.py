from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
import numpy as np

def copytb(ws, start, lengh, start_newtb, border_all):
    for i in range(0, lengh, 1): 
        for j in range(0, lengh, 1):
            col = get_column_letter(j + 1)
            cell = ws[f"{col}{start_newtb+i}"]
            if i == 0 and j == 0:
                cell.value = ""
            else:
                cell.value = f"={col}{i+start}"
            cell.font = Font(bold=False, color="000000", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if i == 0 or j == 0:
                cell.fill = PatternFill("solid", fgColor="5dade2")  
            cell.border = border_all

def colSum(ws, start, lengh, start_newcol, border_all):
    cell = ws[f"A{start_newcol}"]
    cell.value = "SUM"
    cell.font = Font(bold=False, color="000000", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor="f0b27a")  
    cell.border = border_all
    for i in range (0, lengh, 1):
        col = get_column_letter(i + 2)
        cell = ws[f"{col}{start_newcol}"]
        cell.value = f"=SUM({col}{start}:{col}{start+lengh})"
        cell.font = Font(bold=False, color="000000", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="f0b27a")  
        cell.border = border_all

def generate_label(ws, start, lengh, content):
    end_col_letter = get_column_letter(lengh + 1)
    merge_range = f"A{start}:{end_col_letter}{start}"
    ws.merge_cells(merge_range)

    cell = ws[f'A{start}']
    cell.value = content
    cell.font = Font(bold=False, color="FFFFFF", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor="4F81BD")

def generate_tieuchiFull(ws, start, tieuchis, border_all):
    ws[f"A{start}"].fill = PatternFill("solid", fgColor="5dade2")
    ws[f"A{start}"].border = border_all
    for i in range(len(tieuchis)):
        col_letter = get_column_letter(i + 2) 
        cell = ws[f"{col_letter}{start}"]
        cell.value = tieuchis[i]
        cell.font = Font(bold=False, color="000000", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="5dade2")  
        cell.border = border_all
    for i in range(len(tieuchis)):
        cell = ws[f"A{start+i+1}"]
        cell.value = tieuchis[i]
        cell.font = Font(bold=False, color="000000", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="5dade2")
        cell.border = border_all

def generate_phuongsFull(ws, start, phuongs, border_all):
    ws[f"A{start}"].fill = PatternFill("solid", fgColor="5dade2")
    ws[f"A{start}"].border = border_all
    for i in range(len(phuongs)):
        col_letter = get_column_letter(i + 2) 
        cell = ws[f"{col_letter}{start}"]
        cell.value = phuongs[i]
        cell.font = Font(bold=False, color="000000", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="5dade2")  
        cell.border = border_all
    for i in range(len(phuongs)):
        cell = ws[f"A{start+i+1}"]
        cell.value = phuongs[i]
        cell.font = Font(bold=False, color="000000", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="5dade2")
        cell.border = border_all
def school_info_table(ws, start_row, data, border_all):
    headers = ["Trường học", "Ngành học", "Ký hiệu"]
    col_letters = ['A', 'B', 'C']

    for col_idx, header in enumerate(headers):
        cell = ws[f"{col_letters[col_idx]}{start_row}"]
        cell.value = header
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all

    for row_offset, item in enumerate(data, start=1):
        ws[f"A{start_row + row_offset}"].value = item["ten"]
        ws[f"B{start_row + row_offset}"].value = item["nganh"]
        ws[f"C{start_row + row_offset}"].value = item["kihieu"]
        for i in range(3):
            cell = ws[f"{col_letters[i]}{start_row + row_offset}"]
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = border_all

    last_row = start_row + len(data) + 1
    cell = ws[f"D{last_row}"]
    cell.value = "#"
    cell.font = Font(bold=False, color="FFFFFF", size=11)



def generate_chuanhoatb(ws, start, lengh, start_sum_col, start_newtb, border_all):
    for i in range(0, lengh, 1): 
        for j in range(0, lengh, 1):
            col = get_column_letter(i + 2)
            cell = ws[f"{col}{j+start_newtb}"]
            cell.value = f"={col}{j+start}/{col}{start_sum_col}"
            cell.font = Font(bold=False, color="000000", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_all

def generate_col_CW(ws, start, lengh, start_row, start_col, border_all):
    ws[f"{start_col}{start_row}"].fill = PatternFill("solid", fgColor="f4d03f")
    ws[f"{start_col}{start_row}"].border = border_all
    ws[f"{start_col}{start_row}"].value = "Criteria Weight"

    end_quet = get_column_letter(lengh+1)
    for i in range (0, lengh, 1):
        cell = ws[f"{start_col}{i+start_row+1}"]
        cell.value = f"=AVERAGE(B{i+start}:{end_quet}{i+start})"
        cell.font = Font(bold=False, color="000000", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="f4d03f")
        cell.border = border_all

def generate_nhanCW(ws, start, lengh, start_cw_row, start_cw_col, start_newtb, border_all):
    for i in range(0, lengh, 1):
        for j in range(0, lengh, 1):
            col = get_column_letter(i+2)
            cell = ws[f"{col}{j+start_newtb}"]
            cell.value = f"={col}{j+start} * {start_cw_col}{start_cw_row+i}"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_all
    return start_cw_row+lengh, get_column_letter(lengh+2)

def generate_ws(ws, start, lengh, start_new_col, start_new_row, border_all):
    ws[f"{start_new_col}{start_new_row}"].fill = PatternFill("solid", fgColor="f4d03f")
    ws[f"{start_new_col}{start_new_row}"].border = border_all
    ws[f"{start_new_col}{start_new_row}"].value = "Weighted Sum Value"
    for i in range(0, lengh, 1):
        cell = ws[f"{start_new_col}{start_new_row+i+1}"]
        cell.value = f"=SUM(B{start+i}:{get_column_letter(lengh+1)}{start+i})"
        cell.font = Font(bold=False, color="000000", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="f4d03f")
        cell.border = border_all
    col_idx = column_index_from_string(start_new_col) 
    return start_new_row+lengh, get_column_letter(col_idx+1)

def generate_CV(ws, start_col, start_row, lengh, new_col, new_row, border_all):
    ws[f"{new_col}{new_row}"].fill = PatternFill("solid", fgColor="f4d03f")
    ws[f"{new_col}{new_row}"].border = border_all
    ws[f"{new_col}{new_row}"].value = "Consistery Vector"

    start_col_2 = get_column_letter(column_index_from_string(start_col) + 1)
    for i in range(0, lengh, 1):
        cell = ws[f"{new_col}{new_row+i+1}"]
        cell.value = f"={start_col}{start_row+i}/{start_col_2}{start_row+i}"
        cell.font = Font(bold=False, color="000000", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="f4d03f")
        cell.border = border_all

def write_summary(ws, cv_end_row, col_letter, n, border_all):
  
    prev_col_letter = get_column_letter(column_index_from_string(col_letter) - 1)
    row = cv_end_row + 2

    # Lambda max
    cell = ws[f"{prev_col_letter}{row}"]
    cell.value = "Lamda max"
    cell.font = Font(bold=True, color="000000", size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    cell = ws[f"{col_letter}{row}"]
    cv_start_row = cv_end_row - n + 1
    cell.value = f"=AVERAGE({col_letter}{cv_start_row}:{col_letter}{cv_end_row})"
    cell.font = Font(bold=False, color="000000", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_all

    # CI
    row += 1
    cell = ws[f"{prev_col_letter}{row}"]
    cell.value = "CI"
    cell.font = Font(bold=True, color="000000", size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    cell = ws[f"{col_letter}{row}"]
    lamda_cell = f"{col_letter}{row-1}"
    cell.value = f"=({lamda_cell}-{n})/({n}-1)"
    cell.font = Font(bold=False, color="000000", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_all

    # CR
    row += 1
    cell = ws[f"{prev_col_letter}{row}"]
    cell.value = "CR"
    cell.font = Font(bold=True, color="000000", size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    cell = ws[f"{col_letter}{row}"]
    ci_cell = f"{col_letter}{row-1}"
    RI_dict = {
        1: 0.00, 2: 0.00, 3: 0.58, 4: 0.90, 5: 1.12,
        6: 1.24, 7: 1.32, 8: 1.41, 9: 1.45, 10: 1.49
    }
    RI = RI_dict.get(n, 1.49)
    cell.value = f"={ci_cell}/{RI}"
    cell.font = Font(bold=False, color="000000", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_all

def ahpExcel(tieuchis, matran,list_matrices_phuongs, arr_tenNganh):
    n = matran.shape[0]
    wb = Workbook()
    ws = wb.active
    ws.title = "Tính trọng số từng tiêu chí"

    end_col_letter = get_column_letter(len(tieuchis) + 1)
    merge_range = f"A1:{end_col_letter}1"
    ws.merge_cells(merge_range)

    cell = ws['A1']
    cell.value = "Xây dựng ma trận so sánh cặp cho từng tiêu chí"
    cell.font = Font(bold=False, color="FFFFFF", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor="4F81BD")  

    thin = Side(border_style="thin", color="000000")
    border_all = Border(top=thin, left=thin, right=thin, bottom=thin)

    bang1 = 3 
    generate_tieuchiFull(ws, bang1, tieuchis, border_all)
    cell = ws[f'A{bang1}']
    cell.value = "*"
    cell.font = Font(bold=False, color="5dade2", size=11)

    cell = ws[f'{get_column_letter(len(tieuchis) + 2)}{len(tieuchis) + bang1+1}']
    cell.value = "*"
    cell.font = Font(bold=False, color="FFFFFF", size=11)

    for i in range(n):
        for j in range(n):
            cell = ws.cell(row=bang1 + 1 + i, column=2 + j)
            if i == j:
                value = 1.0
            elif i < j:
                value = matran[i][j]
            else:
                col_letter_top = get_column_letter(2 + i)
                row_top = bang1 + 1 + j
                value = f"=1/{col_letter_top}{row_top}"

            cell.value = value
            cell.border = border_all
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # cell.number_format = "??/??" 
             
    tieude = bang1+n+2    
    generate_label(ws, tieude, n, "Tính sum cho từng cột")

    currentcol = tieude+2
    start_bangtinhsum = currentcol
    copytb(ws, bang1, n+1, start_bangtinhsum, border_all)

    currentcol += n+1
    start_col_sum = currentcol
    colSum(ws, bang1+1, n, start_col_sum, border_all)

    currentcol += 2
    generate_label(ws, currentcol, n, "Chuẩn hóa ma trận so sánh cặp")

    currentcol += 2
    generate_tieuchiFull(ws, currentcol, tieuchis, border_all)
    start_tb_chuanhoa = currentcol+1
    generate_chuanhoatb(ws, start_bangtinhsum+1, n, start_col_sum, start_tb_chuanhoa, border_all)

    start_tb_chuanhoa2 = start_tb_chuanhoa + n + 1
    generate_label(ws, start_tb_chuanhoa2, n, "Tính trọng số cho các tiêu chí tính theo từng hàng")
    start_tb_chuanhoa2 += 2
    copytb(ws, start_tb_chuanhoa-1, n+1, start_tb_chuanhoa2, border_all)

    danhdau_tb_ch = start_tb_chuanhoa2+1
    generate_col_CW(ws, danhdau_tb_ch, n, start_tb_chuanhoa2, get_column_letter(n+2), border_all)
    row_Cw = start_tb_chuanhoa2+1
    col_Cw = get_column_letter(n+2)

    start_tb_chuanhoa2 += n+2
    generate_label(ws, start_tb_chuanhoa2, n, "Sử dụng trọng số của các tiêu chí và ma trận so sánh cặp để tính tỷ số nhất quán CR")
    start_tb_chuanhoa2 += 2

    generate_tieuchiFull(ws, start_tb_chuanhoa2, tieuchis, border_all)

    start_tb_chuanhoa2 += 1
    end_row, end_col = generate_nhanCW(ws, bang1+ 1, n, row_Cw, col_Cw, start_tb_chuanhoa2 , border_all)

    end_row_gws, end_col_gws = generate_ws(ws,start_tb_chuanhoa2,n,end_col,start_tb_chuanhoa2-1, border_all )
    generate_col_CW(ws, danhdau_tb_ch, n, start_tb_chuanhoa2-1, end_col_gws, border_all)

    generate_CV(ws, end_col, start_tb_chuanhoa2, n, get_column_letter(column_index_from_string(end_col_gws) + 1), start_tb_chuanhoa2-1, border_all)


    cv_col_letter = get_column_letter(column_index_from_string(end_col_gws) + 1)
    summary_start_row = start_tb_chuanhoa2 + n -1
    write_summary(ws, summary_start_row, cv_col_letter, n, border_all)

    # Sheet 2 
    ws2 = wb.create_sheet(title="Tính trọng số từng phương án")
    def ahpExcel2(ws, tieuchis, phuongs, list_matrices):
        current_row = 1
        thin = Side(border_style="thin", color="000000")
        border_all = Border(top=thin, left=thin, right=thin, bottom=thin)

        school_info_table(ws, current_row, arr_tenNganh, border_all)
        last_row = current_row + len(arr_tenNganh)
        current_row += len(arr_tenNganh) + 3
        cw_row_list = []
        cw_col_list = []

        for idx, tieuchi in enumerate(tieuchis):
            matran = list_matrices[idx]
            n = len(phuongs)

            # Tiêu đề cho mỗi tiêu chí
            end_col_letter = get_column_letter(n + 1)
            ws.merge_cells(f"A{current_row}:{end_col_letter}{current_row}")
            cell = ws[f"A{current_row}"]
            cell.value = f"Ma trận so sánh phương án theo tiêu chí: {tieuchi}"
            cell.font = Font(bold=False, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="4F81BD")

            # Tạo ma trận so sánh phương án
            current_row += 2
            generate_phuongsFull(ws, current_row, phuongs, border_all)
            cell = ws[f"A{current_row}"]
            cell.value = "*"
            cell.font = Font(bold=False, color="5dade2", size=11)


            for i in range(n):
                for j in range(n):
                    cell = ws.cell(row=current_row + 1 + i, column=2 + j)
                    if i == j:
                        cell.value = 1.0
                    elif i < j:
                        cell.value = matran[i][j]
                    else:
                        col_letter_top = get_column_letter(2 + i)
                        row_top = current_row + 1 + j
                        cell.value = f"=1/{col_letter_top}{row_top}"
                    cell.border = border_all
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            cell = ws[f"{get_column_letter(n + 2)}{current_row + n + 1}"]
            cell.value = "*"
            cell.font = Font(bold=False, color="FFFFFF", size=11)

            start_tb = current_row
            current_row += n + 2
            generate_label(ws, current_row, n, "Tính tổng cột")
            current_row += 2
            copytb(ws, start_tb, n+1, current_row, border_all)

            start_bangtinhsum = current_row
            current_row += n + 1
            start_col_sum = current_row
            colSum(ws, start_tb+1, n, start_col_sum, border_all)

            current_row += 2
            generate_label(ws, current_row, n, "Chuẩn hóa ma trận")
            current_row += 2
            generate_phuongsFull(ws, current_row, phuongs, border_all)
            start_chuanhoa = current_row + 1
            generate_chuanhoatb(ws, start_bangtinhsum+1, n, start_col_sum, start_chuanhoa, border_all)

            current_row = start_chuanhoa + n + 2
            generate_label(ws, current_row, n, "Tính trọng số phương án")
            current_row += 2
            copytb(ws, start_chuanhoa-1, n+1, current_row, border_all)
            
            start_tb_chuanhoa2 = current_row

            # Gọi generate_col_CW để tính trọng số phương án
            col_Cw = get_column_letter(n+2)
            generate_col_CW(ws, start_tb_chuanhoa2+1, n, current_row, col_Cw, border_all)

            # Lưu lại đúng dòng/cột CW cho bảng tổng hợp
            cw_row_list.append(current_row + 1)
            cw_col_list.append(col_Cw)

            current_row += n + 2
            generate_label(ws, current_row, n, "Sử dụng trọng số của các tiêu chí và ma trận so sánh cặp để tính tỷ số nhất quán CR")
            current_row += 2

            generate_phuongsFull(ws, current_row, phuongs, border_all)
            current_row += 1

            end_row, end_col = generate_nhanCW(ws, start_tb + 1, n, cw_row_list[-1], col_Cw, current_row, border_all)
            end_row_gws, end_col_gws = generate_ws(ws, current_row, n, end_col, current_row - 1, border_all)

            generate_col_CW(ws, start_tb_chuanhoa2+1, n, current_row - 1, end_col_gws, border_all)

            generate_CV(ws, end_col, current_row, n, get_column_letter(column_index_from_string(end_col_gws) + 1), current_row - 1, border_all)

            cv_col_letter = get_column_letter(column_index_from_string(end_col_gws) + 1)
            summary_start_row = current_row + n - 1
            write_summary(ws, summary_start_row, cv_col_letter, n, border_all)
            current_row += n + 5

        # Hiện bảng CW theo các tiêu chí 
        current_row += 2
        generate_label(ws, current_row, len(phuongs), "Tổng hợp CW của các phương án theo từng tiêu chí")
        current_row += 2
        header_row = current_row
        ws.cell(row=header_row, column=1).fill = PatternFill("solid", fgColor="5dade2")

        for j, tc in enumerate(tieuchis):
            cell = ws.cell(row=header_row, column=2 + j)
            cell.value = tc
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="5dade2")
            cell.border = border_all

        for i, ph in enumerate(phuongs):
            row = header_row + 1 + i
            ws.cell(row=row, column=1).value = ph
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="5dade2")
            ws.cell(row=row, column=1).border = border_all

            for j in range(len(tieuchis)):
                formula = f"={cw_col_list[j]}{cw_row_list[j] + i}"
                cell = ws.cell(row=row, column=2 + j)
                cell.value = formula
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=False, color="000000", size=11)
                cell.border = border_all

        #CW tiêu chí bên sheet 1
        current_row = header_row + len(phuongs) + 1 + 2 
        generate_label(ws, current_row, len(phuongs), "Trọng số tiêu chí")
        ws_weight = wb['Tính trọng số từng tiêu chí']

        for cell in ws_weight[31]:
            if cell.value == "Criteria Weight":
                cw_col_letter = get_column_letter(cell.column)
                break
        cw_start_row = 32

        # Viết dữ liệu
        start_data_row = current_row + 1
        for i, criterion in enumerate(tieuchis):
            row = start_data_row + i

            # Tiêu chí
            cell = ws.cell(row=row, column=1)
            cell.value = criterion
            cell.font = Font(bold=False, color="000000", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="EBF5FB")
            cell.border = border_all

            # Trọng số
            cw_formula = f"='Tính trọng số từng tiêu chí'!{cw_col_letter}{cw_start_row + i}"
            cell = ws.cell(row=row, column=2)
            cell.value = cw_formula
            cell.font = Font(bold=False, color="000000", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="F4D03F")
            cell.border = border_all

        # Bảng Trọng số PA theo các tiêu chí * Trọng số các tiêu chí 
        weight_row_start = start_data_row
        current_row += len(phuongs) + 2
        generate_label(ws, current_row, len(phuongs), "Trọng số PA theo các tiêu chí * Trọng số các tiêu chí ")
        current_row += 2
        header_weighted_row = current_row

        # Header tiêu chí
        ws.cell(row=header_weighted_row, column=1).fill = PatternFill("solid", fgColor="5dade2")
        for j, tc in enumerate(tieuchis):
            cell = ws.cell(row=header_weighted_row, column=2 + j)
            cell.value = tc
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="5dade2")
            cell.border = border_all

        for i, ph in enumerate(phuongs):
            row = header_weighted_row + 1 + i
            
            ws.cell(row=row, column=1).value = ph
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="5dade2")
            ws.cell(row=row, column=1).border = border_all

            for j in range(len(tieuchis)):
                cw_cell_row = header_row + 1 + i
                cw_cell_col = 2 + j
                cw_cell_address = ws.cell(row=cw_cell_row, column=cw_cell_col).coordinate

                weight_cell_row = weight_row_start + j
                weight_cell_col = 2
                weight_cell_address = ws.cell(row=weight_cell_row, column=weight_cell_col).coordinate

                formula = f"={cw_cell_address}*{weight_cell_address}"
                cell = ws.cell(row=row, column=2 + j)
                cell.value = formula
                cell.number_format = "0.000000"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill("solid", fgColor="FCF3CF")
                cell.border = border_all
        #Bảng tổng kết quả 
                total_score_col = 2 + len(tieuchis)
        rank_col = total_score_col + 1

        # Ghi tiêu đề "Tổng điểm" và "Xếp hạng"
        ws.cell(row=header_weighted_row, column=total_score_col).value = "Tổng điểm"
        ws.cell(row=header_weighted_row, column=rank_col).value = "Xếp hạng"
        for col in [total_score_col, rank_col]:
            cell = ws.cell(row=header_weighted_row, column=col)
            cell.fill = PatternFill("solid", fgColor="5dade2")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_all

        # Ghi công thức tính tổng điểm và xếp hạng
        current_row += len(phuongs) + 2
        for i, ph in enumerate(phuongs):
            row = header_weighted_row + 1 + i

            # Tính tổng điểm (SUM hàng hiện tại từ cột 2 đến cột tieuchis+1)
            start_col_letter = get_column_letter(2)
            end_col_letter = get_column_letter(1 + len(tieuchis))
            formula = f"=SUM({start_col_letter}{row}:{end_col_letter}{row})"
            ws.cell(row=row, column=total_score_col).value = formula
            ws.cell(row=row, column=total_score_col).number_format = "0.000000"
            ws.cell(row=row, column=total_score_col).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=total_score_col).fill = PatternFill("solid", fgColor="D6EAF8")
            ws.cell(row=row, column=total_score_col).border = border_all

        # Ghi công thức xếp hạng
        total_start_row = header_weighted_row + 1
        total_end_row = total_start_row + len(phuongs) - 1
        total_col_letter = get_column_letter(total_score_col)

        for i, ph in enumerate(phuongs):
            row = header_weighted_row + 1 + i
            formula_rank = f'=RANK({total_col_letter}{row}, {total_col_letter}{total_start_row}:{total_col_letter}{total_end_row}, 0)'
            cell = ws.cell(row=row, column=rank_col)
            cell.value = formula_rank
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="FADBD8")
            cell.border = border_all

    phuongs = [item['kihieu'] for item in arr_tenNganh]
    ahpExcel2(ws2, tieuchis, phuongs, list_matrices_phuongs)

    
    # wb.save("AHP.xlsx")
    filename = "AHP.xlsx"   
    wb.save(filename)
    return filename

# arr = ["Địa điểm", "Điểm đầu vào", "Học phí", "Chất lượng đào tạo", "Cơ sở vật chất"]
# matrix = np.array([
#     [1, 1/3,1/4, 1/5, 1/3],
#     [3, 1, 3, 3, 2],
#     [4, 1/3, 1, 3, 2],
#     [5, 1/3, 1/3, 1, 1],
#     [3, 1/2, 1/2, 1, 1],
# ])

# arr_tenNganh = [
#     {"ten": "Trường Đại học Công nghệ - Đại học Quốc Gia Hà Nội", "nganh": "Công nghệ thông tin", "kihieu": "QHI-CNTT"},
#     {"ten": "Trường Cao đẳng Kinh tế - Kỹ thuật - Đại học Thái Nguyên", "nganh": "Công nghệ thông tin", "kihieu": "DTU-CNTT"},
#     {"ten": "Trường Đại học Công nghệ thông tin và Truyền thông - Đại học Thái Nguyên", "nganh": "Công nghệ thông tin", "kihieu": "DTC-CNTT"},
#     {"ten": "Trường Đại học Bách Khoa - Đại học Đà Nẵng", "nganh": "Công nghệ thông tin", "kihieu": "DQB-CNTT"},
# ]

# list_matrices_phuongs = [
#     # 1. Địa điểm
#     np.array([
#         [1,   3,   5,   7],
#         [1/3, 1,   2,   4],
#         [1/5, 1/2, 1,   3],
#         [1/7, 1/4, 1/3, 1]
#     ]),
#     # 2. Điểm đầu vào
#     np.array([
#         [1,   2,   4,   6],
#         [1/2, 1,   2,   3],
#         [1/4, 1/2, 1,   2],
#         [1/6, 1/3, 1/2, 1]
#     ]),
#     # 3. Học phí
#     np.array([
#         [1,    1/3, 3,    5],
#         [3,    1,   5,    7],
#         [1/3, 1/5, 1,    3],
#         [1/5, 1/7, 1/3,  1]
#     ]),
#     # 4. Chất lượng đào tạo
#     np.array([
#         [1,     2,     4,     7],
#         [1/2,   1,     3,     5],
#         [1/4, 1/3,     1,     3],
#         [1/7, 1/5, 1/3,     1]
#     ]),
#     # 5. Cơ sở vật chất
#     np.array([
#         [1,   3,   1,   5],
#         [1/3, 1,   1/2, 3],
#         [1,   2,   1,   4],
#         [1/5, 1/3, 1/4, 1]
#     ])
# ]
# ahpExcel(arr, matrix, list_matrices_phuongs, arr_tenNganh)