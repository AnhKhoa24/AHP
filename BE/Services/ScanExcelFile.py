# extractor.py

import openpyxl
from fractions import Fraction

def to_fraction_str(value):
    """
    Chuyển một giá trị (int, float hoặc chuỗi mô tả số) thành chuỗi phân số:
      - Nếu parse thành Fraction thành công và mẫu số != 1 → "tu_so/mau_so"
      - Nếu mẫu số = 1 → trả về chuỗi nguyên, ví dụ "2"
      - Nếu không parse được → trả về "0"
    """
    try:
        frac = Fraction(value).limit_denominator()
    except (ValueError, TypeError):
        return "0"

    if frac.denominator == 1:
        return str(frac.numerator)
    else:
        return f"{frac.numerator}/{frac.denominator}"


def extract_sheet1_from_wb(wb):
    """
    Đọc Sheet1 (index = 0) từ 'wb', tìm cặp dấu '*' (header, footer),
    sau đó:
      - "label": mảng tiêu đề cột (lấy từ hàng header, bỏ ô '*' và ô trống)
      - "full": mảng các dict {"id": <ký tự đầu mỗi từ, in thường>, "text": <label>}
      - "matran": ma trận vuông nxn (n = len(label)), với:
          + nếu i == j → "1"
          + nếu j > i → đọc từ file (hàng header_row_idx+i+1, cột data_column_indices[j])
          + nếu j < i → nghịch đảo (1/x) của giá trị ở (j, i)
    Trả về dict theo mẫu:
    {
      "label": [...],
      "full": [
         {"id": "dd", "text": "Địa điểm"},
         ...
      ],
      "matran": [
         ["1", "2", "4", "5", "1"],
         ["1/2", "1", "2", "3", "2"],
         ...
      ]
    }
    """
    if len(wb.worksheets) < 1:
        raise ValueError("Workbook không có sheet thứ nhất.")
    sheet = wb.worksheets[0]

    # 1) Tìm tất cả các hàng chứa dấu '*'
    marker_rows = []
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if any(cell == "*" for cell in row if cell is not None):
            marker_rows.append(idx)

    if len(marker_rows) < 2:
        raise ValueError("Không tìm thấy đủ 2 dấu '*' trên Sheet1.")

    header_row_idx = marker_rows[0]
    footer_row_idx = marker_rows[1]

    # 2) Xác định "label" và "full"
    header_cells = list(sheet[header_row_idx])
    labels = []
    label_column_indices = []
    for col_idx, cell in enumerate(header_cells, start=1):
        val = cell.value
        if val not in (None, "", "*"):
            labels.append(str(val))
            label_column_indices.append(col_idx)

    # Tạo full = [{"id": ..., "text": ...}, ...]
    full = []
    for label in labels:
        # Mỗi từ lấy ký tự đầu, chuyển xuống lowercase, rồi nối lại
        words = label.split()
        id_str = "".join(word[0].lower() for word in words if len(word) > 0)
        full.append({"id": id_str, "text": label})

    # 3) Xây ma trận vuông nxn
    n = len(labels)
    matran = [["0"] * n for _ in range(n)]

    # Tập hợp danh sách hàng dữ liệu (chỉ những hàng không hoàn toàn trống) từ header_row_idx+1 đến footer_row_idx-1
    row_indices = []
    for r in range(header_row_idx + 1, footer_row_idx):
        row_vals = [cell.value for cell in sheet[r]]
        if all(v in (None, "") for v in row_vals):
            continue
        row_indices.append(r)

    if len(row_indices) < n:
        raise ValueError(f"Số dòng dữ liệu ({len(row_indices)}) không khớp với số label ({n}).")

    # 3.1) Điền giá trị cho từng cặp (i, j)
    for i in range(n):
        for j in range(n):
            if i == j:
                matran[i][j] = "1"
            elif j > i:
                raw = sheet.cell(row=row_indices[i], column=label_column_indices[j]).value
                matran[i][j] = to_fraction_str(raw)
            else:
                prev = matran[j][i]
                try:
                    frac = Fraction(prev)
                    inv = Fraction(frac.denominator, frac.numerator)
                    if inv.denominator == 1:
                        matran[i][j] = str(inv.numerator)
                    else:
                        matran[i][j] = f"{inv.numerator}/{inv.denominator}"
                except Exception:
                    matran[i][j] = "0"

    return {
        "label": labels,
        "full": full,
        "matran": matran
    }


def extract_sheet2_from_wb(wb):
    """
    Đọc Sheet2 (index = 1) từ 'wb'. Có 2 phần chính:

    A) Bảng "ghi chú" ở đầu sheet (từ A2 cho đến khi gặp dấu '#'):
       - Lấy 3 cột: Trường học (A), Ngành học (B), Ký hiệu (C)
       - Dừng khi gặp dòng mà bất kỳ ô nào = '#'
       - Trả về dict "ghichu": {"truong": [...], "nganhhoc": [...], "kihieu": [...]}

    B) Các bảng con đánh dấu bởi cặp dấu '*' giống như Sheet1:
       - Mỗi cặp (*) → một ma trận với text = label hàng (cột A), ma trận vuông tính tương tự Sheet1
       - Gộp thành list "data"

    Trả về dict:
    {
      "ghichu": {
        "truong": [...],
        "nganhhoc": [...],
        "kihieu": [...]
      },
      "data": [
        {
          "text": [...],
          "matran": [...]
        },
        ...
      ]
    }
    """
    if len(wb.worksheets) < 2:
        raise ValueError("Workbook không có sheet thứ hai.")
    sheet = wb.worksheets[1]

    ### A) QUÉT BẢNG GHI CHÚ Ở PHẦN ĐẦU ###
    truong_list = []
    nganhhoc_list = []
    kihieu_list = []

    # Giả sử header của phần ghi chú nằm ở row 1 (A1="Trường học", B1="Ngành học", C1="Ký hiệu").
    # Ta bắt đầu quét từ row = 2 cho đến khi bất kỳ ô nào trong hàng đó = '#'.
    row = 2
    while row <= sheet.max_row:
        # Lấy toàn bộ giá trị của hàng này để kiểm tra xem có '#'
        row_vals = [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]
        # Nếu bất kỳ ô nào trong row_vals có đúng '#', thì dừng ngay
        if any(str(val).strip() == "#" for val in row_vals if val is not None):
            break

        # Nếu không phải '#', ta sẽ lấy 3 cột A, B, C (có thể là rỗng)
        val_a = sheet.cell(row=row, column=1).value  # Cột A
        val_b = sheet.cell(row=row, column=2).value  # Cột B
        val_c = sheet.cell(row=row, column=3).value  # Cột C

        # Nếu cả A, B, C đều trống, bỏ qua (dòng trống xen kẽ)
        if any(v not in (None, "") for v in (val_a, val_b, val_c)):
            truong_list.append(str(val_a) if val_a is not None else "")
            nganhhoc_list.append(str(val_b) if val_b is not None else "")
            kihieu_list.append(str(val_c) if val_c is not None else "")

        row += 1

    ghichu = {
        "truong": truong_list,
        "nganhhoc": nganhhoc_list,
        "kihieu": kihieu_list
    }

    ### B) PHẦN TÌM CÁC BẢNG CON THEO DẤU '*' ###
    # 1) Tìm tất cả các hàng chứa dấu '*'
    marker_rows = []
    for idx, r in enumerate(sheet.iter_rows(values_only=True), start=1):
        if any(cell == "*" for cell in r if cell is not None):
            marker_rows.append(idx)

    if len(marker_rows) % 2 != 0:
        raise ValueError("Số lượng dấu '*' trên Sheet2 không phải số chẵn.")

    results = []
    # 2) Vòng qua từng cặp (header, footer)
    for k in range(0, len(marker_rows), 2):
        header_row_idx = marker_rows[k]
        footer_row_idx = marker_rows[k + 1]

        # 2.1) Xác định các cột dữ liệu: từ hàng header_row_idx, bỏ qua ô '*' và ô trống
        header_cells = list(sheet[header_row_idx])
        data_column_indices = []
        for col_idx, cell in enumerate(header_cells, start=1):
            if cell.value not in (None, "", "*"):
                data_column_indices.append(col_idx)

        # 2.2) Thu danh sách hàng dữ liệu (row_indices) và row_labels
        row_indices = []
        row_labels = []
        for r in range(header_row_idx + 1, footer_row_idx):
            row_vals = [cell.value for cell in sheet[r]]
            if all(v in (None, "") for v in row_vals):
                continue
            row_indices.append(r)
            lbl = sheet.cell(row=r, column=1).value
            row_labels.append(str(lbl) if lbl is not None else "")

        n = len(row_labels)
        if n == 0:
            # Bảng rỗng, bỏ qua
            continue

        if len(data_column_indices) < n:
            raise ValueError(f"Sheet2: Bảng tại header_row {header_row_idx} "
                             f"có {len(data_column_indices)} cột dữ liệu, "
                             f"nhưng cần ít nhất {n} để ghép thành ma trận vuông.")

        # 2.3) Xây ma trận n x n với logic như sheet1
        matran = [["0"] * n for _ in range(n)]
        for i in range(n):
            for j in range(n):
                if i == j:
                    matran[i][j] = "1"
                elif j > i:
                    raw = sheet.cell(row=row_indices[i], column=data_column_indices[j]).value
                    matran[i][j] = to_fraction_str(raw)
                else:
                    prev = matran[j][i]
                    try:
                        frac = Fraction(prev)
                        inv = Fraction(frac.denominator, frac.numerator)
                        if inv.denominator == 1:
                            matran[i][j] = str(inv.numerator)
                        else:
                            matran[i][j] = f"{inv.numerator}/{inv.denominator}"
                    except Exception:
                        matran[i][j] = "0"

        results.append({
            "text": row_labels,
            "matran": matran
        })

    return {
        "ghichu": ghichu,
        "data": results
    }

